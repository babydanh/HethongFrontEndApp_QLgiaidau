#!/usr/bin/env python3
"""
Chạy test → ghi report → update Excel + JSON — 1 lệnh duy nhất.

Cách dùng:
  python scripts/test-and-export.py                    # chạy full
  python scripts/test-and-export.py auth               # chạy 1 module
  python scripts/test-and-export.py all                # chạy full
"""

import json, os, re, subprocess, sys
from datetime import datetime

PROJECT = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
INTEGRATION_DIR = os.path.join(PROJECT, 'integration_test')
REPORT_DIR = os.path.join(PROJECT, 'test_reports')
TC_RE = re.compile(r'(TC-FLUTTER-[A-Z]+-\d+)')

def find_flutter():
    for path in os.environ.get('PATH', '').split(';'):
        for name in ['flutter.bat', 'flutter.exe']:
            if os.path.exists(os.path.join(path, name)):
                return os.path.join(path, name)
    for base in ['C:\\flutter\\bin', 'D:\\flutter\\bin']:
        f = os.path.join(base, 'flutter.bat')
        if os.path.exists(f): return f
    return 'flutter'

def get_test_path(target):
    if not target or target == 'all':
        return INTEGRATION_DIR
    # Map tên ngắn → file test
    mapping = {
        'auth': 'flow_auth_test.dart',
        'home': 'flow_home_test.dart',
        'payment': 'flow_payment_test.dart',
        'notification': 'flow_notification_test.dart',
        'bracket': 'flow_bracket_test.dart',
        'match': 'flow_match_test.dart',
        'profile': 'flow_profile_test.dart',
        'ranking': 'flow_ranking_test.dart',
        'dashboard': 'flow_dashboard_test.dart',
        'intro': 'flow_intro_test.dart',
        'club': 'flow_club_lite_test.dart',
        'team': 'flow_teams_test.dart',
        'admin': 'flow_admin_test.dart',
        'referee': 'flow_referee_test.dart',
        'register': 'flow_register_test.dart',
        'cross': 'flow_cross_test.dart',
        'series': 'flow_series_test.dart',
        'qr': 'flow_qr_test.dart',
        'upload': 'flow_upload_test.dart',
    }
    fname = mapping.get(target, target if target.endswith('.dart') else f'{target}.dart')
    return os.path.join(INTEGRATION_DIR, fname)

def run_test(test_path):
    """Chạy flutter test (có --machine và chỉ định thiết bị) và trả về raw output + test_names."""
    flutter_cmd = find_flutter()
    name = os.path.basename(test_path)
    print(f'\n▶️  Đang chạy: {name}')
    print('⏳  Build...\n')

    process = subprocess.Popen(
        [flutter_cmd, 'test', '--machine', '-d', 'PHM110', test_path],
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        encoding='utf-8', errors='replace', cwd=PROJECT, shell=True
    )

    test_names = {}
    lines = []
    for line in process.stdout:
        line = line.rstrip('\n\r')
        lines.append(line)
        if not line.startswith('{'):
            print(f'  {line}')
        else:
            try:
                ev = json.loads(line)
                if ev.get('type') == 'testStart':
                    # Lưu tên test — test ID nằm trong ev['test']['id']
                    t_obj = ev.get('test', {})
                    t_id = t_obj.get('id') if isinstance(t_obj, dict) else ev.get('testID')
                    t_name = t_obj.get('name', '') if isinstance(t_obj, dict) else ev.get('name', '')
                    test_names[t_id] = t_name
                    print(f'  🔄 {t_name}')
                elif ev.get('type') == 'testDone':
                    # Lấy tên từ mapping đã lưu
                    t_id = ev.get('testID')
                    t_name = test_names.get(t_id, '')
                    icon = '✅' if ev.get('result') == 'success' else '❌'
                    print(f'  {icon} {t_name} ({ev.get("totalTime", 0)}ms)')
            except:
                pass

    process.wait()
    # Lưu raw output để debug
    raw = '\n'.join(lines)
    debug_path = os.path.join(PROJECT, 'test_reports', '_raw_machine.json')
    os.makedirs(os.path.join(PROJECT, 'test_reports'), exist_ok=True)
    with open(debug_path, 'w', encoding='utf-8') as f:
        f.write(raw[:10000])  # Chỉ 10000 ký tự đầu
    print(f'  [DEBUG] Raw machine output saved to {debug_path}')

    return raw, test_names

def parse_results(raw, test_names=None):
    """Parse machine output thành report."""
    if test_names is None:
        test_names = {}

    events = []
    for line in raw.splitlines():
        try:
            parsed = json.loads(line)
            if isinstance(parsed, dict):
                events.append(parsed)
            elif isinstance(parsed, list):
                for item in parsed:
                    if isinstance(item, dict):
                        events.append(item)
        except:
            pass

    # Nếu test_names chưa có, build từ testStart events
    for ev in events:
        if not isinstance(ev, dict): continue
        if ev.get('type') == 'testStart':
            t_obj = ev.get('test')
            t_id = t_obj.get('id') if isinstance(t_obj, dict) else ev.get('testID')
            if t_id not in test_names:
                t_name = ''
                if isinstance(t_obj, dict):
                    t_name = t_obj.get('name', '')
                elif isinstance(t_obj, str):
                    t_name = t_obj
                else:
                    t_name = ev.get('name', '')
                test_names[t_id] = t_name

    results = []
    for ev in events:
        if not isinstance(ev, dict): continue
        if ev.get('type') == 'testDone':
            tid = ev.get('testID')
            t_obj = ev.get('test')
            t_name = ''
            if isinstance(t_obj, dict):
                t_name = t_obj.get('name', '')
            elif isinstance(t_obj, str):
                t_name = t_obj
            name = test_names.get(tid, t_name)
            # Debug: in ra nếu name rỗng
            if not name:
                print(f'  [DEBUG] testDone tid={tid}, t_name="{t_name}", test_names keys={list(test_names.keys())[:5]}')
            m = TC_RE.search(name)
            passed = ev.get('result') == 'success'
            results.append({
                'id': m.group(0) if m else name,
                'name': name,
                'status': 'Pass' if passed else 'Fail',
                'totalTime': ev.get('totalTime', 0),
            })

    return {
        'metadata': {
            'timestamp': datetime.now().isoformat(),
            'total': len(results),
            'passed': sum(1 for r in results if r['status'] == 'Pass'),
            'failed': sum(1 for r in results if r['status'] == 'Fail'),
        },
        'summary': f"{sum(1 for r in results if r['status'] == 'Pass')}/{len(results)} passed",
        'testCases': results,
        'failures': [r for r in results if r['status'] == 'Fail'],
    }

def save_report(report, target):
    os.makedirs(REPORT_DIR, exist_ok=True)
    prefix = target.replace('.dart', '') if target else 'all'

    # Ghi JSON
    json_path = os.path.join(REPORT_DIR, f'{prefix}_report.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f'\n📄  {json_path}')

    # Ghi TXT tóm tắt
    txt_path = os.path.join(REPORT_DIR, f'{prefix}_report.txt')
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write(f"KẾT QUẢ TEST - {report['metadata']['timestamp']}\n")
        f.write(f"{report['summary']}\n\n")
        if report['failures']:
            f.write("--- FAILURES ---\n")
            for r in report['failures']:
                f.write(f"❌ {r['id']}: {r['name']}\n")
        f.write("\n--- ALL ---\n")
        for r in report['testCases']:
            f.write(f"{'✅' if r['status']=='Pass' else '❌'} {r['id']}: {r['status']}\n")
    print(f'📄  {txt_path}')

def update_excel_and_json(report):
    """Cập nhật trực tiếp testcases.json + Excel"""
    # 1. Update testcases.json
    json_path = os.path.join(PROJECT, 'test_docs', 'testcases.json')
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            existing = json.load(f)

        status_map = {r['id']: r['status'] for r in report['testCases']}
        update_count = 0
        for tc in existing:
            if tc['id'] in status_map:
                tc['status'] = status_map[tc['id']]
                update_count += 1

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
        print(f'📄  Updated {json_path} ({update_count} test cases)')

    # 2. Update Excel
    excel_path = os.path.join(PROJECT, 'test_docs', 'testcases.xlsx')
    if os.path.exists(excel_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            update_count = 0
            for row in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=1).value
                if cell_val and str(cell_val).strip() in status_map:
                    ws.cell(row=row, column=9).value = status_map[str(cell_val).strip()]
                    update_count += 1

            wb.save(excel_path)
            print(f'📄  Updated {excel_path} ({update_count} test cases)')
        except ImportError:
            print('⚠️  Không tìm thấy openpyxl, bỏ qua update Excel')
        except Exception as e:
            print(f'⚠️  Lỗi update Excel: {e}')

def main():
    target = ''
    for arg in sys.argv[1:]:
        if not arg.startswith('-'):
            target = arg

    test_path = get_test_path(target)
    name = os.path.basename(test_path).replace('_test.dart', '')

    raw, test_names = run_test(test_path)
    report = parse_results(raw, test_names)
    save_report(report, name)
    update_excel_and_json(report)

    m = report['metadata']
    print(f"\n{'='*50}")
    print(f"  📊  {m['passed']}/{m['total']} passed, {m['failed']} failed")
    print(f"{'='*50}")
    if report['failures']:
        print(f"\n❌ FAILED:")
        for r in report['failures']:
            print(f"   {r['id']}")

if __name__ == '__main__':
    main()
